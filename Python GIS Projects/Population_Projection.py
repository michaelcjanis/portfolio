# -*- coding: utf-8 -*-
"""
Created on Mon Apr  4 08:57:22 2022

@author: Michael.Janis
"""

import pandas as pd
import openpyxl 
import numpy as np
from matplotlib import pyplot as plt
import scipy
from openpyxl import Workbook
import array

wb=openpyxl.load_workbook('File.xlsx')

""" may need to set different sheets active; currently only works for one sheet xlsx"""






sheet=wb.active

data=np.zeros((sheet.max_row-1,sheet.max_column))



for x in range (0, sheet.max_row-1):
    for y in range (0,sheet.max_column):
        data[x,y]=sheet.cell(row=x+2,column=y+1).value
        
        
growth=0.06       
      
is_excess = np.ones(sheet.max_row-1)

pop_proj=np.zeros((sheet.max_row-1, 10))



pop_proj[:,0] = data[:,0] * growth + data[:,0]

    
for j in range (0,10):
    print('iteration')
    excess=2 
    count=0  
    
    """ one iteration: check if projection exceeds max population and if so remove it for redistribution"""
    while  (excess>1):
        count=count+1
       
        excess=0
       
    
        for i in range (0,sheet.max_row-1):
                if (pop_proj [i,j] > data [i,1] ):
                  
                  excess = excess + pop_proj[i,j] - data [i,1]
                  pop_proj [i,j] = data [i,1]
                  is_excess [i] = True 
               
                  
                  
                elif data [i,1] == 0:
                    
                  is_excess [i]=True
                  
                elif data [i,1] == pop_proj [i,j]:
                  is_excess [i]=True  
                  
                else:
                  is_excess [i] = False
      
     
        for k in range (0,sheet.max_row-1):
                if (is_excess [k] == False and data [k,1] > 0) : 
                
                 pop_proj [k,j] = pop_proj [k,j] + (excess / (len(is_excess)-np.sum(is_excess)))
     
    if j<9:
      pop_proj [:,j+1] = pop_proj [:,j] + pop_proj [:,j] * growth

        






wbnew = Workbook()
wbnew_sheet=wbnew.active

for j in range (1,11):
    for a in range (1,sheet.max_row):
        wbnew_sheet.cell(row=a, column =j).value = pop_proj [a-1, j-1]
    


wbnew.save('File.xlsx')
    
    

